from openpyxl import Workbook, load_workbook

wb = load_workbook('C:/Users/user/Desktop/Python/PYTHON_EXCEL/data.xlsx')
# ws = wb.active # Select active work sheet in excel
ws = wb['Sheet1']

print(ws['A1'].value)

# ws['A1'].value = "Hello World"

# ws['A5'].value = '小辉'

# wb.save('C:/Users/user/Desktop/Python/PYTHON_EXCEL/data.xlsx')